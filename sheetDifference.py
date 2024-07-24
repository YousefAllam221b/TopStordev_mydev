import numpy as np
import pandas as pd
import difflib
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import csv
import html
import math
from openpyxl.comments import Comment

def normalize_text(text):
    """Normalize text by converting HTML entities and standardizing line breaks."""
    if isinstance(text, str):
        text = html.unescape(text)  # Convert HTML entities to their corresponding characters
    return text

def read_excel_to_dataframe(file_path, sheet_name):
    """Read the Excel file into a pandas DataFrame without headers."""
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    df.columns = [f'Column{i}' for i in range(1, df.shape[1] + 1)]
    
    # Normalize text in all string columns
    for col in df.columns:
        if df[col].dtype == 'object':  # Check if the column is of string type
            df[col] = df[col].apply(normalize_text)
    
    return df

def detect_bold_column(file_path, sheet_name):
    """Detect the column index that contains the most bold characters, excluding the header row."""
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    bold_counts = [0] * ws.max_column
    
    # Skip the header row (assuming it's the first row)
    for col_idx, col in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
        for cell in col:
            if cell.font.bold:
                bold_counts[col_idx] += 1
    
    # Get the column index with the highest count of bold cells
    max_bold_count = max(bold_counts)
    if max_bold_count == 0:
        return None
    
    bold_column_index = bold_counts.index(max_bold_count)
    return bold_column_index

def dataframe_to_csv_lines(df):
    """Convert a DataFrame to a list of CSV-formatted lines."""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    for row in df.itertuples(index=False, name=None):
        writer.writerow(row)
    return output.getvalue().splitlines()

def compare_lines(old_lines, new_lines):
    """Compare the old and new lines and return a detailed diff."""
    differ = difflib.ndiff(old_lines, new_lines)
    diff = list(differ)
    return diff

def split_into_blocks(df, column_index):
    """Split DataFrame into blocks based on changes in a specific column."""
    blocks = []
    current_block = []
    current_block_value = None
    
    for _, row in df.iterrows():
        # Checks that this is a new block title and it is not nan
        if row.iloc[column_index] != current_block_value and not (isinstance(row.iloc[column_index], float) and math.isnan(row.iloc[column_index])):
            if current_block:
                blocks.append(pd.DataFrame(current_block, columns=df.columns))
            current_block = []
            current_block_value = row.iloc[column_index]
        current_block.append(row)
    if current_block:
        blocks.append(pd.DataFrame(current_block, columns=df.columns))
    
    return blocks

def annotate_changes(diff, df_to_be_written, worksheet):
    """Annotate the DataFrame with changes based on the diff."""
    # Preprocess the diff to create a DataFrame for quick lookup
    diff_df = pd.DataFrame(diff, columns=["line"])
    diff_df["content"] = diff_df["line"].str[2:]
    diff_df["is_deleted"] = diff_df["line"].str.startswith("- ")
    diff_df["is_added"] = diff_df["line"].str.startswith("+ ")
    df2 = diff_df[diff_df['content'].duplicated() == True]
    fill_colors = {
        'Added': PatternFill(start_color='57fa7d', end_color='57fa7d', fill_type='solid'),
        'Deleted': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
        'Other': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    }
    duplicate_content = df2['content'].unique()
    just_added = False
    for _, row in diff_df.iterrows():
        if row["is_added"]:
            temp_df = pd.read_csv(io.StringIO(row['content']), header=None)
            if (just_added):
                # Writing over the deleted row
                df_to_be_written.iloc[df_to_be_written.shape[0] - 1] = temp_df.iloc[0]
                just_added = False
                for cell in worksheet[df_to_be_written.shape[0]]:
                    cell.fill = fill_colors['Other']
            else:
                # Normal Case of adding a new row
                df_to_be_written = pd.concat([df_to_be_written, temp_df], ignore_index=True)
                if row['content'] not in duplicate_content:
                    for cell in worksheet[df_to_be_written.shape[0]]:
                        cell.fill = fill_colors['Added']
            
        elif row['line'].startswith('  '):
            temp_df = pd.read_csv(io.StringIO(row['content']), header=None)
            if (just_added):
                # Writing over the deleted row
                df_to_be_written.iloc[df_to_be_written.shape[0] - 1] = temp_df.iloc[0]
                just_added = False
                for cell in worksheet[df_to_be_written.shape[0]]:
                    cell.fill = fill_colors['Other']
            else:
                # Normal Case of adding a new row
                df_to_be_written = pd.concat([df_to_be_written, temp_df], ignore_index=True)

        elif row["is_deleted"]:
            # If it is a duplicated row, then it is not deleted
            if row['content'] not in duplicate_content:
                temp_df = pd.read_csv(io.StringIO(row['content']), header=None)
                # Adding an empty row to the dataframe that would be overwritten if a new row is added (Added or Other)
                empty_row = len(df_to_be_written.columns) * ""
                df_to_be_written.loc[df_to_be_written.shape[0] + 1] = empty_row
                for col in range(1, len(df_to_be_written.columns) + 1):
                    cell = worksheet.cell(row=df_to_be_written.shape[0], column=col)
                    cell.comment = Comment(str(temp_df.loc[0][col - 1]), "Deleted")
                    cell.fill = fill_colors['Deleted']
                just_added = True
 
    return df_to_be_written

def main(old_file, new_file, output_file):
    old_sheets = pd.ExcelFile(old_file).sheet_names
    new_sheets = pd.ExcelFile(new_file).sheet_names
    sheets = list(set(old_sheets) | set(new_sheets))

    block_column_index = {}
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name in sheets:
            bold_column_index = detect_bold_column(old_file, sheet_name)
            print(f"Bold column index for sheet {sheet_name}: {bold_column_index}")
            if bold_column_index is None:
                print(f"No bold column found in sheet {sheet_name}")
                continue
            block_column_index[sheet_name] = bold_column_index
            
            old_df = read_excel_to_dataframe(old_file, sheet_name)
            new_df = read_excel_to_dataframe(new_file, sheet_name)

            old_blocks = split_into_blocks(old_df, bold_column_index)
            new_blocks = split_into_blocks(new_df, bold_column_index)

            df_to_be_written = pd.DataFrame(columns=[])
            df_to_be_written.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            worksheet = writer.sheets[sheet_name]

            for old_block, new_block in zip(old_blocks, new_blocks):
                old_lines = dataframe_to_csv_lines(old_block)
                new_lines = dataframe_to_csv_lines(new_block)

                diff = compare_lines(old_lines, new_lines)
                df_to_be_written = annotate_changes(diff, df_to_be_written, worksheet)
            df_to_be_written.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

if __name__ == "__main__":
    old_file = 'original.xlsx'
    new_file = 'amended.xlsx'
    output_file = 'output.xlsx'
    main(old_file, new_file, output_file)
